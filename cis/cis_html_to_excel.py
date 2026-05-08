from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from html import unescape
from pathlib import Path
from typing import Iterable, Optional

from bs4 import BeautifulSoup, NavigableString, Tag
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

@dataclass
class ControlBlock:
    control_id: str
    title: str
    full_heading: str
    table: list[list[str]]
    row_count_text: str
    raw_text: str
    query: str = ""
    result: str = ""
    finding_count: int = 0
    expected_output: str = ""
    explanation: str = ""
    remediation: str = ""

HEADER_FILL = PatternFill("solid", fgColor="D9D9A6")  # muted khaki like the screenshot
TEST_FILL = PatternFill("solid", fgColor="F4B183")    # orange highlight
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
BORDER = Border(
    left=Side(style="thin", color="777777"),
    right=Side(style="thin", color="777777"),
    top=Side(style="thin", color="777777"),
    bottom=Side(style="thin", color="777777"),
)
THICK_TOP = Border(top=Side(style="medium", color="777777"))

def clean_text(value: str) -> str:
    value = unescape(value or "")
    value = value.replace("\xa0", " ")
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\s+\n", "\n", value)
    return value.strip()

def get_table_rows(table: Tag) -> list[list[str]]:
    rows: list[list[str]] = []
    for tr in table.find_all("tr"):
        cells = tr.find_all(["th", "td"])
        if not cells:
            continue
        rows.append([clean_text(c.get_text(" ", strip=True)) for c in cells])
    return [r for r in rows if any(c for c in r)]

def iter_until_next_heading(start: Tag) -> Iterable[Tag | NavigableString]:
    for sib in start.next_siblings:
        if isinstance(sib, Tag) and sib.name in {"h2", "h3"}:
            break
        yield sib

def parse_row_count(text: str) -> Optional[int]:
    m = re.search(r"\b(\d+)\s+rows?\s+selected\b", text, flags=re.I)
    if m:
        return int(m.group(1))
    if re.search(r"\bno\s+rows\s+selected\b", text, flags=re.I):
        return 0
    return None

def parse_controls_from_html(html_path: Path) -> tuple[list[list[str]], list[ControlBlock]]:
    soup = BeautifulSoup(html_path.read_text(errors="replace"), "html.parser")

    audit_context: list[list[str]] = []
    for h2 in soup.find_all("h2"):
        if "Audit Context" in h2.get_text(" ", strip=True):
            table = h2.find_next("table")
            if table:
                audit_context = get_table_rows(table)
            break

    controls: list[ControlBlock] = []

    for h3 in soup.find_all("h3"):
        heading = clean_text(h3.get_text(" ", strip=True))
        m = re.match(r"^([0-9]+(?:\.[0-9]+)+)\s*-\s*(.+)$", heading)
        if not m:
            continue

        control_id, title = m.group(1), m.group(2)
        raw_parts: list[str] = []
        table_rows: list[list[str]] = []

        for node in iter_until_next_heading(h3):
            if isinstance(node, NavigableString):
                txt = clean_text(str(node))
                if txt:
                    raw_parts.append(txt)
            elif isinstance(node, Tag):
                if node.name == "table" and not table_rows:
                    table_rows = get_table_rows(node)
                if node.name != "table":
                    txt = clean_text(node.get_text("\n", strip=True))
                    if txt:
                        raw_parts.append(txt)

        raw_text = "\n".join(raw_parts)
        row_count = parse_row_count(raw_text)
        if row_count is None:
            row_count = max(len(table_rows) - 1, 0) if table_rows else 0

        controls.append(
            ControlBlock(
                control_id=control_id,
                title=title,
                full_heading=heading,
                table=table_rows,
                row_count_text=raw_text,
                raw_text=raw_text,
                finding_count=row_count,
            )
        )

    return audit_context, controls

def parse_queries_from_sql_runner(sql_path: Optional[Path]) -> dict[str, str]:
    if not sql_path or not sql_path.exists():
        return {}

    text = sql_path.read_text(errors="replace")
    queries: dict[str, str] = {}

    pattern = re.compile(
        r"PROMPT\s+<h3>\s*([0-9]+(?:\.[0-9]+)+)\s*-\s*.*?</h3>\s*"
        r"PROMPT\s+<p>Finding rows:</p>\s*"
        r"(?P<sql>.*?);(?=\s*PROMPT\s+<h3>|\s*PROMPT\s+<h2>Cleanup</h2>|\Z)",
        flags=re.I | re.S,
    )

    for m in pattern.finditer(text):
        cid = m.group(1)
        sql = m.group("sql").strip()
        sql = re.sub(r"\n{3,}", "\n\n", sql)
        queries[cid] = sql

    return queries

def expected_output_for(control: ControlBlock) -> str:
    title = control.title.upper()

    if "(MANUAL)" in title:
        return "Manual review required; returned rows are evidence for review, not an automatic failure."

    return "Empty result / no rows selected."

def remediation_for(control: ControlBlock) -> str:
    cid = control.control_id
    title = control.title.upper()

    remediations = {
        "2.3.1": "ALTER SYSTEM SET BACKGROUND_CORE_DUMP = 'partial' SCOPE = SPFILE;",
        "2.3.2": "ALTER SYSTEM SET SHADOW_CORE_DUMP = 'partial' SCOPE = SPFILE;",
        "2.3.3": "ALTER SYSTEM SET ALLOW_GROUP_ACCESS_TO_SGA = FALSE SCOPE = SPFILE;",
        "2.3.5": "ALTER SYSTEM SET OS_ROLES = FALSE SCOPE = SPFILE;",
        "2.3.6": "ALTER SYSTEM SET REMOTE_OS_ROLES = FALSE SCOPE = SPFILE;",
        "2.3.7": "ALTER SYSTEM SET SEC_MAX_FAILED_LOGIN_ATTEMPTS = 3 SCOPE = SPFILE;",
        "2.3.8": "ALTER SYSTEM SET SEC_PROTOCOL_ERROR_FURTHER_ACTION = '(DROP,3)' SCOPE = SPFILE;",
        "2.3.9": "ALTER SYSTEM SET SEC_PROTOCOL_ERROR_TRACE_ACTION = LOG SCOPE = SPFILE;",
        "2.3.10": "ALTER SYSTEM SET SEC_RETURN_SERVER_RELEASE_BANNER = FALSE SCOPE = SPFILE;",
        "2.3.11": "ALTER SYSTEM SET REMOTE_LOGIN_PASSWORDFILE = 'NONE' SCOPE = SPFILE;",
        "2.3.12": "ALTER SYSTEM SET REMOTE_LISTENER = '' SCOPE = SPFILE;",
        "2.3.13": "ALTER SYSTEM SET RESOURCE_LIMIT = TRUE SCOPE = SPFILE;",
        "2.3.14": "ALTER SYSTEM SET REMOTE_OS_AUTHENT = FALSE SCOPE = SPFILE;",
        "2.3.15": "ALTER SYSTEM SET SEC_CASE_SENSITIVE_LOGON = TRUE SCOPE = SPFILE;",
        "3.1": "ALTER PROFILE <profile_name> LIMIT FAILED_LOGIN_ATTEMPTS 5;",
        "3.2": "ALTER PROFILE <profile_name> LIMIT PASSWORD_LOCK_TIME 1;",
        "3.3": "ALTER PROFILE <profile_name> LIMIT PASSWORD_LIFE_TIME <value> PASSWORD_GRACE_TIME <value>; ensure the combined total is <= 365.",
        "3.4": "ALTER PROFILE <profile_name> LIMIT PASSWORD_REUSE_MAX UNLIMITED;",
        "3.5": "ALTER PROFILE <profile_name> LIMIT PASSWORD_VERIFY_FUNCTION <approved_verify_function>;",
        "3.6": "Review the password verify function source and update it to meet organisational password complexity policy; use an approved custom function or Oracle strong verify function where appropriate.",
        "3.7": "ALTER PROFILE <profile_name> LIMIT PASSWORD_ROLLOVER_TIME 0;",
        "3.8": "ALTER PROFILE <profile_name> LIMIT INACTIVE_ACCOUNT_TIME 120;",
        "4.4": "Expire/reset passwords so only current password versions are used.",
        "4.7": "DROP PUBLIC DATABASE LINK <link_name>;",
    }

    if cid in remediations:
        return remediations[cid]

    if "REVOKED" in title or "REVOKE" in title:
        return "Revoke the listed privilege/role/object grant from unauthorized grantees after confirming business need."
    if "AUDIT" in title and "(AUTOMATED)" in title:
        return "Create or enable the required audit policy/action according to the benchmark remediation guidance."
    if "PASSWORD" in title:
        return "Alter the affected profile(s) to match the CIS benchmark requirement."
    if "MANUAL" in title:
        return "Review the returned evidence and remediate according to the benchmark and organisational policy."
    return "Review the returned rows and apply the CIS benchmark remediation for this control."

def explain_failure(control: ControlBlock) -> str:
    title = control.title.upper()
    n = control.finding_count

    if "(MANUAL)" in title:
        return (
            f"This is a manual CIS control. The audit returned {n} row(s) of evidence that "
            "must be reviewed to determine whether the configuration complies with policy."
        )

    if n == 0:
        return "No findings were returned by the audit query."

    if "PASSWORD_LIFE_TIME" in title and "PASSWORD_GRACE_TIME" in title:
        return (
            f"The audit returned {n} profile row(s) where PASSWORD_LIFE_TIME plus "
            "PASSWORD_GRACE_TIME exceeds the CIS threshold."
        )
    if "PASSWORD_LIFE_TIME" in title:
        return f"The audit returned {n} profile row(s) where PASSWORD_LIFE_TIME exceeds the CIS threshold."
    if "PASSWORD_REUSE_MAX" in title:
        return (
            f"The audit returned {n} profile row(s) where PASSWORD_REUSE_MAX is not set "
            "to the CIS-required value."
        )
    if "FAILED_LOGIN_ATTEMPTS" in title:
        return (
            f"The audit returned {n} profile row(s) where FAILED_LOGIN_ATTEMPTS exceeds "
            "the CIS maximum or is set to UNLIMITED."
        )
    if "REMOTE_LOGIN_PASSWORDFILE" in title:
        return (
            f"The audit returned {n} row(s) showing REMOTE_LOGIN_PASSWORDFILE is not set "
            "to NONE."
        )
    if "SEC_PROTOCOL_ERROR_TRACE_ACTION" in title:
        return (
            f"The audit returned {n} row(s) showing SEC_PROTOCOL_ERROR_TRACE_ACTION is "
            "not set to LOG."
        )
    if "RESOURCE_LIMIT" in title:
        return f"The audit returned {n} row(s) showing RESOURCE_LIMIT is not set to TRUE."
    if "OS_ROLES" in title or "REMOTE_OS_ROLES" in title or "REMOTE_OS_AUTHENT" in title:
        return f"The audit returned {n} row(s) showing this parameter is not set to FALSE."
    if "REVOKED" in title or "REVOKE" in title:
        return f"The audit returned {n} unauthorized or review-required grant/privilege row(s)."
    if "AUDIT" in title:
        return f"The audit returned {n} row(s), indicating the expected audit action/policy is not enabled as required."

    return (
        f"The audit query returned {n} row(s). For this automated CIS check, returned rows "
        "represent findings because the expected output is an empty result set."
    )

def classify_result(control: ControlBlock, include_manual_review_as_finding: bool = True) -> str:
    title = control.title.upper()
    if "(MANUAL)" in title:
        return "REVIEW" if include_manual_review_as_finding else "MANUAL"
    return "FAIL" if control.finding_count > 0 else "PASS"

def write_wrapped_cell(ws, row: int, col: int, value: str, bold: bool = False, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font = Font(bold=bold)
    if fill:
        cell.fill = fill
    return cell

def apply_table_style(ws, start_row: int, start_col: int, rows: list[list[str]]):
    if not rows:
        return

    max_cols = max(len(r) for r in rows)
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx in range(start_col, start_col + max_cols):
            value = row[c_idx - start_col] if c_idx - start_col < len(row) else ""
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r_idx == start_row:
                cell.fill = HEADER_FILL
                cell.font = Font(bold=True, color="0070C0")
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

def write_appendix_workbook(
    audit_context: list[list[str]],
    controls: list[ControlBlock],
    xlsx_path: Path,
    include_passes: bool = False,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "CIS Benchmark Output"

    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A1"

    row = 1
    write_wrapped_cell(ws, row, 1, "Appendix III: CIS Benchmark Output", bold=True)
    ws.cell(row=row, column=1).font = Font(bold=True, underline="single", size=12)
    row += 2

    if audit_context:
        write_wrapped_cell(ws, row, 1, "[Audit Context]", bold=True, fill=SECTION_FILL)
        row += 1
        apply_table_style(ws, row, 1, audit_context)
        row += len(audit_context) + 2

    figure_no = 1

    for control in controls:
        control.result = classify_result(control)
        control.expected_output = expected_output_for(control)
        control.explanation = explain_failure(control)
        control.remediation = remediation_for(control)

        if not include_passes and control.result == "PASS":
            continue

        ws.cell(row=row, column=1, value="-" * 55)
        ws.cell(row=row, column=1).border = THICK_TOP
        row += 1

        test_cell = write_wrapped_cell(ws, row, 1, f"[Test]: {control.title}", bold=True, fill=TEST_FILL)
        row += 1

        query_text = control.query or "Not captured in HTML output. Re-run converter with --sql-script to include the audit SQL."
        write_wrapped_cell(ws, row, 1, f"[Query]: {query_text}", bold=False)
        row += 1

        write_wrapped_cell(ws, row, 1, "[Output]:", bold=True)
        row += 1

        if control.table:
            table_start = row
            apply_table_style(ws, table_start, 1, control.table)
            row += len(control.table)
            if control.finding_count:
                ws.cell(row=row, column=1, value=f"{control.finding_count} row(s) selected.")
                row += 1
            clean_title = re.sub(r"\s*\((Automated|Manual)\)\s*$", "", control.title, flags=re.I)
            caption = f"Figure {figure_no}: Output returned for {control.control_id} - {clean_title}."
            write_wrapped_cell(ws, row, 1, caption, bold=False)
            ws.cell(row=row, column=1).font = Font(italic=True)
            figure_no += 1
            row += 1
        else:
            ws.cell(row=row, column=1, value="No rows selected.")
            row += 1

        write_wrapped_cell(ws, row, 1, f"[Expected Output]: {control.expected_output}", bold=False)
        row += 1

        label = "[Why Non-Compliant]" if control.result == "FAIL" else "[Review Required]" if control.result == "REVIEW" else "[Result]"
        write_wrapped_cell(ws, row, 1, f"{label}: {control.explanation}", bold=False)
        row += 1

        write_wrapped_cell(ws, row, 1, f"[Remediation]: {control.remediation}", bold=False)
        row += 2

    widths = {
        "A": 34, "B": 42, "C": 32, "D": 32, "E": 24, "F": 24, "G": 24, "H": 24,
        "I": 24, "J": 24, "K": 24, "L": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for cells in ws.iter_rows():
        ws.row_dimensions[cells[0].row].height = None
        for cell in cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if isinstance(val, str) and val.startswith("[Query]:"):
            ws.row_dimensions[r].height = 60

    wb.save(xlsx_path)

def main() -> None:
    parser = argparse.ArgumentParser(description="Convert CIS SQL*Plus HTML audit output into Excel appendix format.")
    parser.add_argument("html", type=Path, help="SQL*Plus HTML audit output file.")
    parser.add_argument("-o", "--output", type=Path, default=Path("CIS_Benchmark_Output.xlsx"), help="Output .xlsx path.")
    parser.add_argument("--sql-script", type=Path, default=None, help="Optional audit runner .sql file, used to populate [Query] text.")
    parser.add_argument("--include-passes", action="store_true", help="Include PASS/no-row controls as well as findings/review items.")
    args = parser.parse_args()

    audit_context, controls = parse_controls_from_html(args.html)
    queries = parse_queries_from_sql_runner(args.sql_script)

    for control in controls:
        control.query = queries.get(control.control_id, "")

    if not controls:
        raise SystemExit("No <h3> audit controls found in the HTML file. Check the input file.")

    write_appendix_workbook(audit_context, controls, args.output, include_passes=args.include_passes)

    fail_count = sum(1 for c in controls if classify_result(c) == "FAIL")
    review_count = sum(1 for c in controls if classify_result(c) == "REVIEW")
    pass_count = sum(1 for c in controls if classify_result(c) == "PASS")

    print(f"Created: {args.output}")
    print(f"Controls parsed: {len(controls)}")
    print(f"FAIL: {fail_count} | REVIEW: {review_count} | PASS: {pass_count}")
    if not args.sql_script:
        print("Tip: pass --sql-script oracle_19c_cis_database_audit_runner_v2.sql to populate [Query] lines.")

if __name__ == "__main__":
    main()
